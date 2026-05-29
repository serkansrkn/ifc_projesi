# ifc_pipeline/exporter.py
"""
Normalize edilmis DataFrame'i Excel ve JSON'a aktarir.
Sheet 1: Tam liste
Sheet 2: Tip Bazli Detayli Metraj (Malzeme/Tip kırılımı ile)
Sheet 3: Maliyet Ozeti
Sheet 4: QA raporu
"""
from __future__ import annotations
import json
import logging
import pandas as pd
import numpy as np
from typing import Optional

try:
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
    HAS_STYLES = True
except ImportError:
    HAS_STYLES = False

logger = logging.getLogger(__name__)

# ─── Excel Stil Tanımları ────────────────────────────────────────────────────
# Tüm stil nesneleri MODÜL SEVİYESİNDE bir kez oluşturulur — hücre başına
# yeni nesne oluşturulmaz. openpyxl her atamada zaten nesneyi kopyalar.

if HAS_STYLES:
    HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    HEADER_FONT = Font(bold=True, color="FFFFFF", size=11, name="Calibri")
    HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
    THIN_BORDER = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )
    DATA_FONT = Font(name="Calibri", size=10)  # Tek seferlik oluştur
    ALT_ROW_FILL = PatternFill(start_color="F2F7FB", end_color="F2F7FB", fill_type="solid")
    NUMBER_FORMAT_2 = "#,##0.00"

# Hücre-bazlı formatlama eşiği — bu satır sayısının üzerinde
# sadece başlık formatlanır, veri satırları stilsiz bırakılır.
# 50.000 satır × 18 sütun = 900.000 hücre işlemi → saatlerce sürer.
MAX_ROWS_FOR_CELL_STYLING = 5000


def to_excel(df, output_path, quality=None, include_pivot=True):
    """Ana Excel export fonksiyonu — çok sekmeli, formatlanmış rapor."""
    logger.info("Excel yazılıyor: %s (%d satır)", output_path, len(df))

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        export_df = _prep(df)
        
        # 1. Ana Ham Veri (Tam Liste)
        export_df.to_excel(writer, sheet_name="Element Listesi", index=False)
        _autowidth(writer, "Element Listesi", export_df)
        _format_sheet(writer, "Element Listesi", export_df)

        if include_pivot and not df.empty:
            # 2. Tip ve Malzeme Kırılımlı Detaylı Metraj
            detailed_summ = _detailed_type_summary(df)
            if not detailed_summ.empty:
                detailed_summ.to_excel(writer, sheet_name="Detayli Metraj", index=False)
                _autowidth(writer, "Detayli Metraj", detailed_summ)
                _format_sheet(writer, "Detayli Metraj", detailed_summ)

            # 3. Maliyet Özeti (Eğer maliyet hesaplanmışsa)
            cost_cols = [c for c in df.columns if c.startswith("cost_")]
            if cost_cols:
                cost = _cost_summary(df, cost_cols[0])
                cost.to_excel(writer, sheet_name="Maliyet Ozeti", index=False)
                _autowidth(writer, "Maliyet Ozeti", cost)
                _format_sheet(writer, "Maliyet Ozeti", cost)

        # 4. Veri Kalitesi ve Uyarılar
        if quality:
            qa = _qa_df(quality)
            qa.to_excel(writer, sheet_name="Veri Kalitesi", index=False)
            _autowidth(writer, "Veri Kalitesi", qa)
            _format_sheet(writer, "Veri Kalitesi", qa)

    logger.info("Excel yazma tamamlandı: %s", output_path)


def _prep(df):
    """Excel'e yazmadan önce boolean ve float değerleri temizler."""
    df = df.copy()
    for col in df.columns:
        if pd.api.types.is_bool_dtype(df[col]) or str(df[col].dtype) == "boolean":
            df[col] = df[col].map({True: "Evet", False: "Hayir", pd.NA: ""})
        elif hasattr(df[col], "cat"):
            df[col] = df[col].astype(str)
    
    # Sayısal alanları temizle (virgülden sonra çok uzamasını engelle)
    for col in df.select_dtypes(include="float").columns:
        df[col] = df[col].round(4)
    return df


def _autowidth(writer, sheet_name, df):
    """Excel sütun genişliklerini içeriğe göre otomatik ayarlar."""
    try:
        ws = writer.sheets[sheet_name]

        for i, col in enumerate(df.columns, 1):
            header_len = len(str(col))

            if df.empty:
                mx = header_len
            else:
                # Büyük DataFrame'lerde sample al — tamamını taramak yavaş
                sample = df[col].head(1000) if len(df) > 1000 else df[col]
                try:
                    max_data_len = sample.astype(str).str.len().max()
                except Exception:
                    max_data_len = 10
                mx = max(header_len, int(max_data_len) if pd.notna(max_data_len) else header_len)

            ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = min(mx + 4, 60)
    except Exception as e:
        logger.debug("Sütun genişliği ayarlanamadı (%s): %s", sheet_name, e)


def _format_sheet(writer, sheet_name, df):
    """
    Excel sayfasına profesyonel formatlama uygular.
    
    Performans koruması: Veri satırı sayısı MAX_ROWS_FOR_CELL_STYLING'i aşarsa
    sadece başlık satırı formatlanır. Aksi halde hücre-bazlı işlem saatlerce sürer.
    """
    if not HAS_STYLES:
        return
    try:
        ws = writer.sheets[sheet_name]
        row_count = ws.max_row - 1  # Başlık hariç veri satırı sayısı
        
        # ── Başlık satırı (her zaman formatlanır) ─────────────────────────
        for cell in ws[1]:
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = HEADER_ALIGNMENT
            cell.border = THIN_BORDER

        # Başlık satırını sabitle
        ws.freeze_panes = "A2"

        # ── Veri satırları (sadece küçük tablolarda) ──────────────────────
        if row_count > MAX_ROWS_FOR_CELL_STYLING:
            logger.info(
                "%s: %d satır var — hücre formatlama atlanıyor (eşik: %d). "
                "Sadece başlık formatlandı.",
                sheet_name, row_count, MAX_ROWS_FOR_CELL_STYLING
            )
            return

        # Sayısal sütun indekslerini önceden tespit et
        float_col_indices = set()
        for i, col in enumerate(df.columns):
            if df[col].dtype in ("float64", "float32"):
                float_col_indices.add(i)

        for row_idx in range(2, ws.max_row + 1):
            is_alt_row = (row_idx % 2 == 0)
            for col_idx, cell in enumerate(ws[row_idx]):
                cell.border = THIN_BORDER
                cell.font = DATA_FONT  # Modül seviyesinde oluşturulmuş nesne
                # Sayısal sütunlara format uygula
                if col_idx in float_col_indices:
                    cell.number_format = NUMBER_FORMAT_2
                # Zebra renklendirme
                if is_alt_row:
                    cell.fill = ALT_ROW_FILL

    except Exception as e:
        logger.warning("Sayfa formatlanamadı (%s): %s", sheet_name, e)


def _detailed_type_summary(df):
    """
    Sadece 'wall' veya 'column' demek yerine, Tip Adı ve Malzeme kırılımı 
    yaparak alt toplamları verir (Gerçek metraj mantığı).
    """
    groups = []
    
    # Sadece gerekli sütunları kopyala (tüm DataFrame'i kopyalamak yerine)
    needed_cols = ["element_type", "type_name", "materials", "area_m2", "volume_m3", "length_m"]
    existing_cols = [c for c in needed_cols if c in df.columns]
    df_copy = df[existing_cols].copy()

    df_copy["type_name"] = df_copy["type_name"].fillna("Belirsiz Tip")
    if "materials" not in df_copy.columns:
        df_copy["materials"] = "Belirsiz Malzeme"
    else:
        df_copy["materials"] = df_copy["materials"].fillna("Belirsiz Malzeme")
        df_copy.loc[df_copy["materials"] == "", "materials"] = "Belirsiz Malzeme"

    # Eleman Tipi -> Malzeme -> Tip Adı hiyerarşisinde grupla
    for (elem_type, material, type_name), sub in df_copy.groupby(
        ["element_type", "materials", "type_name"], observed=True
    ):
        if sub.empty:
            continue
            
        groups.append({
            "Kategori": str(elem_type).upper(),
            "Malzeme (Material)": str(material),
            "Tip Adı (Type)": str(type_name),
            "Adet": len(sub),
            "Alan (m2)": round(sub["area_m2"].sum(skipna=True), 2) if ("area_m2" in sub.columns and sub["area_m2"].notna().any()) else None,
            "Hacim (m3)": round(sub["volume_m3"].sum(skipna=True), 2) if ("volume_m3" in sub.columns and sub["volume_m3"].notna().any()) else None,
            "Uzunluk (m)": round(sub["length_m"].sum(skipna=True), 2) if ("length_m" in sub.columns and sub["length_m"].notna().any()) else None,
        })
        
    result = pd.DataFrame(groups)
    if not result.empty:
        result = result.sort_values(by=["Kategori", "Malzeme (Material)"])
    return result


def _cost_summary(df, cost_col="cost_TL"):
    """Maliyetleri detaylı şekilde özetler."""
    groups = []
    # Sadece gerekli sütunları kopyala
    needed = ["element_type", "type_name", cost_col]
    if "unit_price" in df.columns:
        needed.append("unit_price")
    existing = [c for c in needed if c in df.columns]
    df_copy = df[existing].copy()
    df_copy["type_name"] = df_copy["type_name"].fillna("Belirsiz Tip")

    for (et, type_name), sub in df_copy.groupby(["element_type", "type_name"], observed=True):
        if cost_col not in sub.columns:
            continue
        total = sub[cost_col].sum(skipna=True)
        if total > 0:
            groups.append({
                "Element Tipi": str(et).upper(),
                "Tip Adı": str(type_name),
                "Adet": len(sub),
                "Birim Fiyat": sub["unit_price"].iloc[0] if "unit_price" in sub.columns else None,
                f"Toplam Maliyet ({cost_col.split('_')[-1]})": round(total, 2),
            })
            
    result = pd.DataFrame(groups)
    if not result.empty:
        cost_label = f"Toplam Maliyet ({cost_col.split('_')[-1]})"
        result.loc[len(result)] = {
            "Element Tipi": "GENEL", 
            "Tip Adı": "TOPLAM", 
            "Adet": result["Adet"].sum(),
            "Birim Fiyat": None,
            cost_label: result[cost_label].sum()
        }
    return result


def _qa_df(quality):
    """Veri kalitesi özetini tabloya dönüştürür."""
    rows = [{"Metrik": "Toplam Element", "Deger": quality.get("total", 0), "Detay": ""}]
    for et, cnt in quality.get("by_type", {}).items():
        rows.append({"Metrik": f"  {et}", "Deger": cnt, "Detay": "adet"})
    rows.append({"Metrik": "", "Deger": "", "Detay": ""})
    rows.append({"Metrik": "Metraj Kapsami", "Deger": "", "Detay": ""})
    for et, pct in quality.get("coverage", {}).items():
        rows.append({"Metrik": f"  {et}", "Deger": f"%{pct}", "Detay": ""})
    rows.append({"Metrik": "", "Deger": "", "Detay": ""})
    rows.append({"Metrik": "Uyarilar", "Deger": "", "Detay": ""})
    for w in quality.get("warnings", []):
        rows.append({"Metrik": "", "Deger": "", "Detay": w})
    return pd.DataFrame(rows)


def to_json(df, output_path, orient="records"):
    """DataFrame'i JSON formatında dışa aktarır."""
    # NaN → None, boş string → None (tutarlılık)
    clean = df.copy()
    for col in clean.select_dtypes(include="float").columns:
        clean[col] = clean[col].round(6)
    for col in clean.select_dtypes(include="object").columns:
        clean[col] = clean[col].replace("", None)
    clean = clean.where(clean.notna(), other=None)
    records = clean.to_dict(orient=orient)
    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(records, f, ensure_ascii=False, indent=2, default=str)