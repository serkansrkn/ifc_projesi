#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
ifc_pipeline/exporter.py modülü için unit testler.
Excel ve JSON export, formatlama ve yardımcı fonksiyon testleri.
"""
import pytest
import sys
import os
import io
import json
import tempfile
import numpy as np
import pandas as pd

sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

from ifc_pipeline.exporter import (
    to_excel,
    to_json,
    _prep,
    _detailed_type_summary,
    _cost_summary,
    _qa_df,
)


# ─── Yardımcı fonksiyonlar ───────────────────────────────────────────────────

def _make_df(n=5, **overrides):
    """Test DataFrame oluşturur."""
    base = {
        "global_id": [f"id_{i}" for i in range(n)],
        "element_type": pd.Categorical(["wall"] * n),
        "ifc_class": ["IfcWall"] * n,
        "name": [f"Duvar-{i}" for i in range(n)],
        "type_name": ["Beton Duvar 200mm"] * n,
        "level": ["1. Kat"] * n,
        "level_elevation": [3.0] * n,
        "phase": [""] * n,
        "is_external": pd.array([True] * n, dtype="boolean"),
        "load_bearing": pd.array([False] * n, dtype="boolean"),
        "fire_rating": ["F90"] * n,
        "area_m2": [15.5] * n,
        "volume_m3": [3.1] * n,
        "length_m": [5.0] * n,
        "thickness_m": [0.2] * n,
        "materials": ["C30/37"] * n,
        "source_software": pd.Categorical(["revit"] * n),
        "source_file": ["test.ifc"] * n,
    }
    base.update(overrides)
    return pd.DataFrame(base)


# ─── _prep testleri ───────────────────────────────────────────────────────────

class TestPrep:

    def test_boolean_to_evet_hayir(self):
        df = _make_df(1)
        result = _prep(df)
        assert result.iloc[0]["is_external"] == "Evet"
        assert result.iloc[0]["load_bearing"] == "Hayir"

    def test_float_rounding(self):
        df = _make_df(1, volume_m3=[0.123456789012345])
        result = _prep(df)
        # volume_m3 → 9 basamak
        assert result.iloc[0]["volume_m3"] == round(0.123456789012345, 9)

    def test_area_rounding(self):
        df = _make_df(1, area_m2=[15.123456789])
        result = _prep(df)
        # area_m2 → 6 basamak
        assert result.iloc[0]["area_m2"] == round(15.123456789, 6)

    def test_category_to_string(self):
        df = _make_df(1)
        result = _prep(df)
        # pandas 3'te category → StringDtype olabilir, object de olabilir
        assert result["element_type"].dtype != "category"

    def test_na_boolean(self):
        df = _make_df(1, is_external=pd.array([pd.NA], dtype="boolean"))
        result = _prep(df)
        assert result.iloc[0]["is_external"] == ""


# ─── to_excel testleri ────────────────────────────────────────────────────────

class TestToExcel:

    def test_basic_excel_output(self, tmp_path):
        df = _make_df(3)
        output = tmp_path / "test.xlsx"
        to_excel(df, str(output))
        assert output.exists()
        assert output.stat().st_size > 0

    def test_excel_has_sheets(self, tmp_path):
        df = _make_df(3)
        output = tmp_path / "test.xlsx"
        qa = {"total": 3, "by_type": {"wall": 3}, "coverage": {"wall": 100.0}, "warnings": []}
        to_excel(df, str(output), quality=qa)

        import openpyxl
        wb = openpyxl.load_workbook(str(output))
        sheet_names = wb.sheetnames
        assert "Element Listesi" in sheet_names
        assert "Detayli Metraj" in sheet_names
        assert "Veri Kalitesi" in sheet_names

    def test_excel_with_bytesio(self):
        """BytesIO buffer'a yazabilmeli."""
        df = _make_df(2)
        buffer = io.BytesIO()
        to_excel(df, buffer)
        buffer.seek(0)
        assert len(buffer.read()) > 0

    def test_empty_df(self, tmp_path):
        df = _make_df(0)
        output = tmp_path / "empty.xlsx"
        to_excel(df, str(output))
        assert output.exists()

    def test_no_pivot(self, tmp_path):
        df = _make_df(3)
        output = tmp_path / "nopivot.xlsx"
        to_excel(df, str(output), include_pivot=False)

        import openpyxl
        wb = openpyxl.load_workbook(str(output))
        assert "Element Listesi" in wb.sheetnames
        assert "Detayli Metraj" not in wb.sheetnames


# ─── to_json testleri ─────────────────────────────────────────────────────────

class TestToJson:

    def test_basic_json_output(self, tmp_path):
        df = _make_df(3)
        output = tmp_path / "test.json"
        to_json(df, str(output))
        assert output.exists()

        with open(str(output), encoding="utf-8") as f:
            data = json.load(f)
        assert len(data) == 3

    def test_json_nan_becomes_null(self, tmp_path):
        df = _make_df(1, area_m2=[np.nan])
        output = tmp_path / "nan.json"
        to_json(df, str(output))

        with open(str(output), encoding="utf-8") as f:
            content = f.read()
        data = json.loads(content)
        # NaN → null (JSON) veya NaN (pandas serialization farkı)
        assert data[0]["area_m2"] is None or (isinstance(data[0]["area_m2"], float) and np.isnan(data[0]["area_m2"]))

    def test_json_empty_string_becomes_null(self, tmp_path):
        df = _make_df(1, materials=[""])
        output = tmp_path / "empty_str.json"
        to_json(df, str(output))

        with open(str(output), encoding="utf-8") as f:
            content = f.read()
        data = json.loads(content)
        # Boş string → None veya NaN
        val = data[0]["materials"]
        assert val is None or (isinstance(val, float) and np.isnan(val))

    def test_json_precision(self, tmp_path):
        df = _make_df(1, volume_m3=[0.123456789012345])
        output = tmp_path / "precision.json"
        to_json(df, str(output))

        with open(str(output), encoding="utf-8") as f:
            data = json.load(f)
        # 9 basamak
        assert data[0]["volume_m3"] == round(0.123456789012345, 9)


# ─── _detailed_type_summary testleri ──────────────────────────────────────────

class TestDetailedTypeSummary:

    def test_basic_summary(self):
        df = _make_df(5)
        result = _detailed_type_summary(df)
        assert not result.empty
        assert "Kategori" in result.columns
        assert "Malzeme (Material)" in result.columns
        assert "Tip Adı (Type)" in result.columns

    def test_missing_materials_handled(self):
        df = _make_df(3, materials=["", "", ""])
        result = _detailed_type_summary(df)
        assert (result["Malzeme (Material)"] == "Belirsiz Malzeme").all()

    def test_none_type_name(self):
        df = _make_df(2, type_name=[None, None])
        result = _detailed_type_summary(df)
        assert (result["Tip Adı (Type)"] == "Belirsiz Tip").all()

    def test_grouping(self):
        df = pd.concat([
            _make_df(2, type_name=["Tip A"] * 2, materials=["Beton"] * 2),
            _make_df(3, type_name=["Tip B"] * 3, materials=["Çelik"] * 3,
                     global_id=[f"id_extra_{i}" for i in range(3)]),
        ], ignore_index=True)
        result = _detailed_type_summary(df)
        assert len(result) == 2  # 2 ayrı grup

    def test_empty_df(self):
        df = _make_df(0)
        result = _detailed_type_summary(df)
        assert result.empty


# ─── _cost_summary testleri ───────────────────────────────────────────────────

class TestCostSummary:

    def test_basic_cost(self):
        df = _make_df(3, unit_price=[100.0] * 3)
        df["cost_TL"] = df["area_m2"] * df["unit_price"]
        result = _cost_summary(df, "cost_TL")
        assert not result.empty
        # Son satır TOPLAM olmalı
        assert result.iloc[-1]["Element Tipi"] == "GENEL"
        assert result.iloc[-1]["Tip Adı"] == "TOPLAM"

    def test_no_cost_column(self):
        df = _make_df(2)
        result = _cost_summary(df, "cost_TL")
        assert result.empty

    def test_total_row_sum(self):
        df = _make_df(2, unit_price=[100.0] * 2)
        df["cost_TL"] = [1000.0, 2000.0]
        result = _cost_summary(df, "cost_TL")
        total_row = result[result["Element Tipi"] == "GENEL"]
        assert total_row.iloc[0]["Toplam Maliyet (TL)"] == 3000.0


# ─── _qa_df testleri ──────────────────────────────────────────────────────────

class TestQaDf:

    def test_basic_qa(self):
        quality = {
            "total": 10,
            "by_type": {"wall": 7, "slab": 3},
            "coverage": {"wall": 85.0, "slab": 100.0},
            "warnings": ["Test uyarısı"],
        }
        result = _qa_df(quality)
        assert not result.empty
        assert "Metrik" in result.columns
        assert "Deger" in result.columns

    def test_empty_quality(self):
        quality = {"total": 0, "by_type": {}, "coverage": {}, "warnings": []}
        result = _qa_df(quality)
        assert not result.empty

    def test_warnings_in_output(self):
        quality = {
            "total": 5,
            "by_type": {},
            "coverage": {},
            "warnings": ["Uyarı 1", "Uyarı 2"],
        }
        result = _qa_df(quality)
        details = result["Detay"].tolist()
        assert "Uyarı 1" in details
        assert "Uyarı 2" in details


if __name__ == "__main__":
    pytest.main([__file__, "-v"])
